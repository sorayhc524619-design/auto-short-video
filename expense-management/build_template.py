"""
副業経費管理キット ひな形生成スクリプト v0.1
9シートを作成し、タブ色・タイトル・ヘッダーを設定する。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 色定義(デザインルールより)
NAVY = "1F4E79"
WHITE = "FFFFFF"
LIGHT_YELLOW = "FFF8DC"  # 入力可能セル背景
LIGHT_RED_BG = "F8D7DA"  # 免責バナー背景
DARK_RED = "C00000"      # 警告文字
GRAY_TEXT = "404040"     # 通常文字

# 共通スタイル生成関数
def header_font():
    return Font(name="メイリオ", size=11, bold=True, color=WHITE)

def header_fill():
    return PatternFill("solid", fgColor=NAVY)

def body_font():
    return Font(name="メイリオ", size=10, color=GRAY_TEXT)

def input_fill():
    return PatternFill("solid", fgColor=LIGHT_YELLOW)

def thin_border():
    side = Side(border_style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

# シート定義(シート名, タブ色, タイトル)
SHEETS = [
    ("00_スタート", "1F4E79", "副業経費管理キット v1.0"),
    ("01_入力", "5B9BD5", "レシート入力"),
    ("02_科目マスタ", "A9D18E", "勘定科目マスタ"),
    ("03_店舗辞書", "FFD966", "店舗名→勘定科目 辞書"),
    ("04_ダッシュボード", "ED7D31", "経費ダッシュボード"),
    ("05_科目別集計", "C00000", "勘定科目別 年間集計"),
    ("06_20万円判定", "F4B183", "雑所得20万円 判定シート"),
    ("07_設定", "A6A6A6", "基本設定"),
    ("99_使い方", "B4A7D6", "使い方ガイド"),
]

def apply_header_row(ws, row, start_col_idx, headers, widths):
    """指定行にヘッダー行を配置し、列幅を設定する"""
    for i, (text, width) in enumerate(zip(headers, widths)):
        col_idx = start_col_idx + i
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        # 列幅設定(openpyxlでは get_column_letter が必要)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 28


def setup_sheet_01_input(ws):
    """01_入力シートのヘッダー設定"""
    headers = ["日付", "金額(税込)", "店舗名", "勘定科目", "メモ", "支払方法", "家事按分率", "按分後金額", "重複フラグ"]
    widths = [12, 13, 25, 15, 30, 13, 10, 13, 10]
    # 入力シートは特例で2行目ヘッダー・3行目データ(今回は見出しのみ)
    apply_header_row(ws, 4, 2, headers, widths)


def setup_sheet_02_master(ws):
    """02_科目マスタシートのヘッダー設定"""
    headers = ["科目コード", "科目名", "副業での該当例", "家事按分の目安", "デフォルト按分率", "備考"]
    widths = [10, 15, 45, 10, 12, 40]
    apply_header_row(ws, 4, 2, headers, widths)


def setup_sheet_03_dict(ws):
    """03_店舗辞書シートのヘッダー設定"""
    headers = ["店舗名(部分一致キー)", "推奨勘定科目", "信頼度", "備考"]
    widths = [25, 15, 8, 30]
    apply_header_row(ws, 4, 2, headers, widths)


def setup_sheet_05_summary(ws):
    """05_科目別集計シートのヘッダー設定"""
    headers = ["科目名", "按分前 年間合計", "按分後 年間合計", "按分により減額", "確定申告書の記入欄", "記入時のコツ"]
    widths = [15, 15, 15, 15, 30, 40]
    apply_header_row(ws, 4, 2, headers, widths)


def place_disclaimer_banner(ws, start_row, end_col_letter="J"):
    """免責バナー(薄赤背景+濃い赤文字)を指定行から4行分配置"""
    from openpyxl.utils import get_column_letter
    banner_text = (
        "⚠ 本ツールの判定は参考情報です。住民税・医療費控除・給与2,000万円超の場合など例外があります。"
        "最終判断は税務署または税理士にご確認をお願いします。ツールの判定結果に対する責任は負いかねます。"
    )
    # B列〜end_col_letter列をマージして配置
    merge_range = f"B{start_row}:{end_col_letter}{start_row + 2}"
    ws.merge_cells(merge_range)
    cell = ws.cell(row=start_row, column=2, value=banner_text)
    cell.font = Font(name="メイリオ", size=10, bold=True, color=DARK_RED)
    cell.fill = PatternFill("solid", fgColor=LIGHT_RED_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(start_row, start_row + 3):
        ws.row_dimensions[r].height = 22


def setup_sheet_00_start(ws):
    """00_スタートシート:タグライン+免責バナー"""
    ws["B3"] = "〜 副業会社員のための、ぜんぶ入り経費&申告サポート 〜"
    ws["B3"].font = Font(name="メイリオ", size=10, color=GRAY_TEXT)
    # サマリーカード配置エリアのプレースホルダ(枠のみ)
    place_disclaimer_banner(ws, start_row=18)


def setup_sheet_06_handan(ws):
    """06_20万円判定シート:免責バナー+月別収入入力枠"""
    place_disclaimer_banner(ws, start_row=4)
    # 月別入力ヘッダー(B8:C8)
    ws["B8"] = "月"
    ws["C8"] = "副業収入(円)"
    for c in ("B8", "C8"):
        ws[c].font = header_font()
        ws[c].fill = header_fill()
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[8].height = 28
    # 1月〜12月の行
    from openpyxl.utils import get_column_letter
    for i in range(12):
        r = 9 + i
        ws.cell(row=r, column=2, value=f"{i+1}月").font = body_font()
        income = ws.cell(row=r, column=3, value=0)
        income.font = body_font()
        income.fill = input_fill()
        income.number_format = '#,##0"円"'


def setup_sheet_99_howto(ws):
    """99_使い方シート:免責バナーのみ最下部に配置"""
    ws["B4"] = "このシートには使い方ガイドを掲載します(M3で作成)。"
    ws["B4"].font = body_font()
    place_disclaimer_banner(ws, start_row=40)


def setup_sheet_07_config(ws):
    """07_設定シートの項目配置"""
    items = [
        ("ユーザー名", ""),
        ("対象年度(西暦)", 2026),
        ("デフォルト家事按分率", "100%"),
        ("前年度副業所得(比較用)", 0),
        ("マクロ動作モード", "標準(自動サジェスト有効)"),
        ("", ""),
        ("ファイルバージョン", "v1.0"),
    ]
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    for i, (label, default) in enumerate(items):
        r = 4 + i
        ws.cell(row=r, column=2, value=label).font = Font(name="メイリオ", size=10, bold=True, color=NAVY)
        c = ws.cell(row=r, column=3, value=default)
        c.font = body_font()
        if label and label != "ファイルバージョン":
            c.fill = input_fill()


def build():
    wb = Workbook()
    # デフォルトで1枚作られるシートを削除
    wb.remove(wb.active)

    for name, tab_color, title in SHEETS:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = tab_color

        # A列を余白列として幅2に
        ws.column_dimensions["A"].width = 2

        # タイトルをB2セルに配置
        ws["B2"] = title
        ws["B2"].font = Font(name="メイリオ", size=16, bold=True, color=NAVY)

        # 行の高さを設定
        ws.row_dimensions[1].height = 10
        ws.row_dimensions[2].height = 36

        # シートごとの固有セットアップ
        if name == "00_スタート":
            setup_sheet_00_start(ws)
        elif name == "01_入力":
            setup_sheet_01_input(ws)
        elif name == "02_科目マスタ":
            setup_sheet_02_master(ws)
        elif name == "03_店舗辞書":
            setup_sheet_03_dict(ws)
        elif name == "05_科目別集計":
            setup_sheet_05_summary(ws)
        elif name == "06_20万円判定":
            setup_sheet_06_handan(ws)
        elif name == "07_設定":
            setup_sheet_07_config(ws)
        elif name == "99_使い方":
            setup_sheet_99_howto(ws)

    # 00_スタートを先頭表示にする
    wb.active = 0

    output = "副業経費管理キット_v0.1.xlsx"
    wb.save(output)
    print(f"生成完了: {output}")

if __name__ == "__main__":
    build()
